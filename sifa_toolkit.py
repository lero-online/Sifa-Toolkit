# -*- coding: utf-8 -*-
"""
SiFa Toolkit – Alles für die Fachkraft für Arbeitssicherheit
Robuster App-Rahmen:
- Page-Config früh
- Session-Init sicher
- main() Wrapper + sichtbare Fehleranzeige
- Keine harten st.stop() – Seite bleibt bedienbar
"""

import traceback
import json
from dataclasses import dataclass, asdict, field
from datetime import date, datetime
from typing import List, Optional, Dict, Any, Tuple
from dateutil.relativedelta import relativedelta
from io import BytesIO
import re

import pandas as pd
import streamlit as st
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule


# --- Seite früh konfigurieren ---
st.set_page_config(
    page_title="SiFa Toolkit – Alles für die Fachkraft für Arbeitssicherheit",
    layout="wide"
)

# =========================
# Datenmodelle
# =========================

STOP_LEVELS = [
    "S (Substitution/Quelle entfernen)",
    "T (Technisch)",
    "O (Organisatorisch)",
    "P (PSA)",
    "Q (Qualifikation/Unterweisung)"
]
STATUS_LIST = ["offen", "in Umsetzung", "wirksam", "nicht wirksam", "entfallen"]

@dataclass
class Measure:
    title: str
    stop_level: str
    responsible: str = ""
    due_date: Optional[str] = None  # ISO
    status: str = "offen"
    notes: str = ""

@dataclass
class Hazard:
    id: str
    area: str
    activity: str
    hazard: str
    sources: List[str]
    existing_controls: List[str]
    prob: int = 3
    sev: int = 3
    risk_value: int = 9
    risk_level: str = "mittel"
    additional_measures: List[Measure] = field(default_factory=list)
    last_review: Optional[str] = None
    reviewer: str = ""
    documentation_note: str = ""

@dataclass
class Assessment:
    company: str
    location: str
    created_at: str
    created_by: str
    industry: str = "Hotel/Gastgewerbe"
    scope_note: str = ""
    risk_matrix_thresholds: Dict[str, List[int]] = field(default_factory=lambda: {"thresholds": [6, 12, 16]})
    hazards: List[Hazard] = field(default_factory=list)
    measures_plan_note: str = ""
    documentation_note: str = ""
    next_review_hint: str = ""

# =========================
# Utility
# =========================

def compute_risk(prob: int, sev: int, thresholds: List[int]) -> Tuple[int, str]:
    v = prob * sev
    if v <= thresholds[0]:
        return v, "niedrig"
    elif v <= thresholds[1]:
        return v, "mittel"
    elif v <= thresholds[2]:
        return v, "hoch"
    else:
        return v, "sehr hoch"

def hazard_to_row(h: Hazard) -> Dict[str, Any]:
    return {
        "ID": h.id, "Bereich": h.area, "Tätigkeit": h.activity, "Gefährdung": h.hazard,
        "Quellen/Einwirkungen": "; ".join(h.sources), "Bestehende Maßnahmen": "; ".join(h.existing_controls),
        "Eintrittswahrscheinlichkeit (1-5)": h.prob, "Schadensschwere (1-5)": h.sev,
        "Risikosumme": h.risk_value, "Risikostufe": h.risk_level,
        "Letzte Prüfung": h.last_review or "", "Prüfer/in": h.reviewer,
        "Dokumentationshinweis": h.documentation_note
    }

def measures_to_rows(h: Hazard) -> List[Dict[str, Any]]:
    rows = []
    for m in h.additional_measures:
        rows.append({
            "Gefährdungs-ID": h.id, "Bereich": h.area, "Gefährdung": h.hazard,
            "Maßnahme": m.title, "STOP(+Q)": m.stop_level, "Verantwortlich": m.responsible,
            "Fällig am": m.due_date or "", "Status": m.status, "Hinweis": m.notes
        })
    return rows

def new_id(prefix="HZ", n=4) -> str:
    ts = datetime.now().strftime("%y%m%d%H%M%S%f")[-n:]
    return f"{prefix}-{int(datetime.now().timestamp())}-{ts}"

def dump_excel(assess: Assessment) -> bytes:
    # --- Datenaufbereitung ---
    hazards_df = pd.DataFrame([hazard_to_row(h) for h in assess.hazards])
    measures_df = pd.DataFrame([r for h in assess.hazards for r in measures_to_rows(h)])

    # Maßnahmen-Plan (Schritt 5) – inkl. Status/Verantwortlich/Fällig
    plan_rows = []
    for h in assess.hazards:
        for m in h.additional_measures:
            plan_rows.append({
                "Gefährdungs-ID": h.id,
                "Bereich": h.area,
                "Tätigkeit": h.activity,
                "Gefährdung": h.hazard,
                "Risikosumme": h.risk_value,
                "Risikostufe": h.risk_level,
                "Maßnahme": m.title,
                "STOP(+Q)": m.stop_level,
                "Verantwortlich": m.responsible,
                "Fällig am": m.due_date or "",
                "Status": m.status,
                "Hinweis": m.notes,
            })
    plan_df = pd.DataFrame(plan_rows)

    # Wirksamkeit (Schritt 6) je Gefährdung
    review_rows = []
    for h in assess.hazards:
        review_rows.append({
            "Gefährdungs-ID": h.id,
            "Bereich": h.area,
            "Tätigkeit": h.activity,
            "Gefährdung": h.hazard,
            "Letzte Prüfung": h.last_review or "",
            "Prüfer/in": h.reviewer,
            "Beurteilungs-/Dokumentationshinweis": h.documentation_note,
        })
    review_df = pd.DataFrame(review_rows)

    # Meta / Stammdaten (Schritt 1)
    meta = {
        "Unternehmen": assess.company,
        "Standort": assess.location,
        "Erstellt am": assess.created_at,
        "Erstellt von": assess.created_by,
        "Branche": assess.industry,
        "Umfang/Scope": assess.scope_note,
    }
    meta_df = pd.DataFrame(list(meta.items()), columns=["Feld", "Wert"])

    # Dokumentation (Schritt 7)
    doc_df = pd.DataFrame({"Dokumentationshinweis": [assess.documentation_note or ""]})

    # Fortschreiben (Schritt 8)
    prog_df = pd.DataFrame({"Anlässe/Fristen (Fortschreibung)": [assess.next_review_hint or ""]})

    # Konfiguration
    thresholds = assess.risk_matrix_thresholds.get("thresholds", [6, 12, 16])
    conf_df = pd.DataFrame(
        {"Einstellung": ["Grenze niedrig (≤)", "Grenze mittel (≤)", "Grenze hoch (≤)"],
         "Wert": [thresholds[0], thresholds[1], thresholds[2]]}
    )

    # --- Excel schreiben ---
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        # Reihenfolge/Blätter:
        meta_df.to_excel(writer, sheet_name="01_Stammdaten", index=False)
        hazards_df.to_excel(writer, sheet_name="10_Gefährdungen", index=False)
        measures_df.to_excel(writer, sheet_name="20_Maßnahmen", index=False)
        plan_df.to_excel(writer, sheet_name="30_Plan", index=False)
        review_df.to_excel(writer, sheet_name="40_Wirksamkeit", index=False)
        doc_df.to_excel(writer, sheet_name="50_Dokumentation", index=False)
        prog_df.to_excel(writer, sheet_name="60_Fortschreiben", index=False)
        conf_df.to_excel(writer, sheet_name="90_Konfiguration", index=False)

        wb = writer.book

        # Styling Helper
        header_fill = PatternFill("solid", fgColor="E6EEF8")
        bold = Font(bold=True)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_sheet(name: str, freeze: bool = True, wide_wrap: bool = True):
            ws = wb[name]
            # Überschriften-Format
            if ws.max_row >= 1:
                for c in ws[1]:
                    c.font = bold
                    c.fill = header_fill
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = border

            # Inhalte
            if ws.max_row >= 2 and ws.max_column >= 1:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if wide_wrap:
                            try:
                                cell.alignment = cell.alignment.copy(horizontal="left", vertical="top", wrap_text=True)
                            except Exception:
                                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        cell.border = border

            # Spaltenbreiten (autofit grob, limitiert)
            for col_idx in range(1, ws.max_column + 1):
                col = get_column_letter(col_idx)
                maxlen = 8
                limit = min(ws.max_row, 200)  # Performance
                for r in range(1, limit + 1):
                    val = ws.cell(row=r, column=col_idx).value
                    if val is None:
                        continue
                    maxlen = max(maxlen, len(str(val)))
                ws.column_dimensions[col].width = min(maxlen + 2, 60)

            # Freeze Pane
            if freeze and ws.max_row > 1:
                ws.freeze_panes = "A2"
            return ws

        # Stil auf alle relevanten Blätter
        for sheet in ["01_Stammdaten", "10_Gefährdungen", "20_Maßnahmen", "30_Plan",
                      "40_Wirksamkeit", "50_Dokumentation", "60_Fortschreiben",
                      "90_Konfiguration"]:
            wide = sheet not in ["01_Stammdaten", "90_Konfiguration"]
            style_sheet(sheet, freeze=True, wide_wrap=wide)

        # Dropdown für Status im Plan-Blatt
        if "30_Plan" in wb.sheetnames:
            ws_plan = wb["30_Plan"]
            if ws_plan.max_row >= 2 and ws_plan.max_column >= 1:
                # Finde Spalte "Status"
                status_col_idx = None
                for c in range(1, ws_plan.max_column + 1):
                    if (ws_plan.cell(row=1, column=c).value or "").strip() == "Status":
                        status_col_idx = c
                        break
                if status_col_idx:
                    dv = DataValidation(
                        type="list",
                        formula1='"' + ",".join(STATUS_LIST) + '"',
                        allow_blank=True,
                        showDropDown=True,
                    )
                    ws_plan.add_data_validation(dv)
                    dv.add(f"{get_column_letter(status_col_idx)}2:{get_column_letter(status_col_idx)}1048576")

        # Farbskala (Risiko-Ampel) im Gefährdungsblatt auf "Risikosumme"
        if "10_Gefährdungen" in wb.sheetnames:
            ws_h = wb["10_Gefährdungen"]
            # Spalte "Risikosumme" suchen
            risk_col = None
            for c in range(1, ws_h.max_column + 1):
                if (ws_h.cell(row=1, column=c).value or "").strip() == "Risikosumme":
                    risk_col = c
                    break
            if risk_col:
                # 3-Farbskala: grün -> gelb -> rot
                col_letter = get_column_letter(risk_col)
                rng = f"{col_letter}2:{col_letter}{ws_h.max_row}"
                rule = ColorScaleRule(
                    start_type="num", start_value=1, start_color="C6EFCE",
                    mid_type="num", mid_value=max(2, thresholds[1]), mid_color="FFEB9C",
                    end_type="num", end_value=max(3, thresholds[2] + 1), end_color="F8CBAD"
                )
                ws_h.conditional_formatting.add(rng, rule)

        # Druckfreundliche Kopfzeile (einfach)
        for name in wb.sheetnames:
            ws = wb[name]
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0  # beliebig viele Seiten in der Höhe

    bio.seek(0)
    return bio.read()


def as_json(assess: Assessment) -> str:
    return json.dumps(asdict(assess), ensure_ascii=False, indent=2)

def from_json(s: str) -> Assessment:
    data = json.loads(s)
    hazards = []
    for h in data.get("hazards", []):
        measures = [Measure(**m) for m in h.get("additional_measures", [])]
        hazards.append(Hazard(
            id=h["id"], area=h["area"], activity=h["activity"], hazard=h["hazard"],
            sources=h.get("sources", []),
            existing_controls=h.get("existing_controls", h.get("existing", [])),
            prob=h.get("prob", 3), sev=h.get("sev", 3),
            risk_value=h.get("risk_value", 9), risk_level=h.get("risk_level", "mittel"),
            additional_measures=measures, last_review=h.get("last_review"),
            reviewer=h.get("reviewer", ""), documentation_note=h.get("documentation_note", "")
        ))
    return Assessment(
        company=data.get("company",""), location=data.get("location",""),
        created_at=data.get("created_at",""), created_by=data.get("created_by",""),
        industry=data.get("industry","Hotel/Gastgewerbe"), scope_note=data.get("scope_note", ""),
        risk_matrix_thresholds=data.get("risk_matrix_thresholds", {"thresholds":[6,12,16]}),
        hazards=hazards, measures_plan_note=data.get("measures_plan_note",""),
        documentation_note=data.get("documentation_note",""), next_review_hint=data.get("next_review_hint","")
    )

def slug(*parts: str) -> str:
    s = "_".join(parts)
    s = re.sub(r"[^a-zA-Z0-9_-]+", "_", s)
    return s[:80]

# ===== Splitting für Mehrfach-Gefährdungen =====

_SPLIT_PATTERN = re.compile(r"\s*(?:,|/| und | & )\s*")

def split_hazard_text(text: str) -> List[str]:
    """Teilt 'Gefährdung' auf: Trennzeichen Komma, Slash, 'und', '&'."""
    if not text:
        return []
    parts = [p.strip() for p in _SPLIT_PATTERN.split(text) if p and p.strip()]
    # Duplikate eliminieren, Reihenfolge beibehalten
    seen, uniq = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p)
            uniq.append(p)
    return uniq or [text.strip()]

# =========================
# Branchen-Bibliothek (ERWEITERT)
# =========================

def M(title, stop="O (Organisatorisch)"):
    return {"title": title, "stop_level": stop}

# --- HOTEL/GAST ---
LIB_HOTEL = {
    "Küche": [
        {"activity": "Kochen (Töpfe/Kessel)", "hazard": "Hitze, heiße Flüssigkeiten, Verbrühungen/Verbrennungen", "sources": ["Herde","Kessel","Töpfe"], "existing": ["Hitzeschutz"], "measures":[M("Topfdeckel/Spritzschutz nutzen","T (Technisch)"), M("‚Heiß!‘ rufen"), M("Hitzeschutzhandschuhe","P (PSA)")]},
        {"activity": "Braten (Pfanne/Grillplatte)", "hazard": "Fettspritzer, Verbrennungen, Rauch/Dämpfe", "sources": ["Pfannen","Grillplatten"], "existing": ["Abzug"], "measures":[M("Spritzschutz einsetzen","T (Technisch)"), M("Haube reinigen/prüfen")]},
        {"activity": "Frittieren", "hazard": "Fettbrand, Verbrennungen, Spritzer", "sources": ["Fritteusen"], "existing": ["Fettbrandlöscher"], "measures":[M("Ölwechsel-/Reinigungsplan"), M("Hitzeschutzschürze & Handschuhe","P (PSA)")]},
        {"activity": "Kombidämpfer öffnen", "hazard": "Dampf/Heißluft – Verbrühung beim Öffnen", "sources": ["Kombidämpfer"], "existing": ["Abkühlzeit"], "measures":[M("Tür erst spaltweise öffnen"), M("Hitzeschutzhandschuhe","P (PSA)")]},
        {"activity": "Saucen/Reduktionen", "hazard": "Dampf, Spritzer, inhalative Belastung", "sources": ["Reduktion"], "existing": ["Abluft"], "measures":[M("Deckel/Spritzschutz","T (Technisch)"), M("Lüftung checken")]},
        {"activity": "Schneiden mit Messern", "hazard": "Schnitt-/Stichverletzungen", "sources": ["Messer"], "existing": ["Scharfe Messer"], "measures":[M("Schleifplan"), M("Schnittschutzhandschuhe bei Bedarf","P (PSA)")]},
        {"activity": "Aufschnittmaschine", "hazard": "Schnittverletzungen an rotierenden Klingen", "sources": ["Aufschnitt"], "existing": ["Schutzhaube","Not-Aus"], "measures":[M("Sicherheitsbauteile prüfen","T (Technisch)"), M("Nur befugte Bedienung")]},
        {"activity": "Fleischwolf/Gemüseschneider", "hazard": "Eingezogenwerden, Schnittverletzung", "sources": ["Wolf","Gemüseschneider"], "existing": ["Stopfer"], "measures":[M("Stopfer verwenden"), M("Unterweisung Not-Aus","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Kippkessel/Bräter", "hazard": "Verbrühung, Quetschen beim Kippen", "sources": ["Kippkessel"], "existing": ["Hitzeschutz"], "measures":[M("Kipp-Prozess standardisieren"), M("Zweihandbedienung beachten","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Spülbereich", "hazard": "Heißes Wasser/Dampf, Chemikalien, Rutschgefahr", "sources": ["Spülmaschine","Klarspüler"], "existing": ["Hand-/Augenschutz"], "measures":[M("Sofort-Wisch-Regel"), M("Antirutsch-Matten","T (Technisch)")]},
        {"activity": "Reinigung/Chemie", "hazard": "Ätz-/Reizwirkung, Chlorgas bei Mischungen", "sources": ["Reiniger/Desinfektion"], "existing": ["Dosiersysteme"], "measures":[M("Vordosierte Kartuschen","S (Substitution/Quelle entfernen)"), M("Betriebsanweisungen aushängen")]},
        {"activity": "Gasgeräte", "hazard": "Gasleck, CO-Bildung, Brand/Explosion", "sources": ["Gasherde","Leitungen"], "existing": ["Dichtheitsprüfung"], "measures":[M("Gaswarnmelder","T (Technisch)"), M("Leckcheck vor Inbetriebnahme")]},
        {"activity": "Warenannahme/Hubwagen", "hazard": "Quetschungen, Heben/Tragen, Verkehrswege", "sources": ["Rollcontainer","Hubwagen"], "existing": ["Hebehilfen"], "measures":[M("Wege kennzeichnen"), M("Kurzunterweisung Heben/Tragen","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Altöl/Müll entsorgen", "hazard": "Verbrennung bei heißem Öl, Schnitt/Infektion", "sources": ["Altöl","Müllsack"], "existing": ["Abkühlen"], "measures":[M("Deckel-Transportbehälter","T (Technisch)"), M("Handschutz verpflichtend","P (PSA)")]},
        {"activity": "TK-/Kühlräume", "hazard": "Kälte, Rutschgefahr, Einsperr-Risiko", "sources": ["Kühlzelle","TK"], "existing": ["Kälteschutz"], "measures":[M("Tür-Notöffnung prüfen","T (Technisch)"), M("Aufenthaltsdauer begrenzen")]},
        {"activity": "Allergenmanagement", "hazard": "Kreuzkontamination/Allergene", "sources": ["Zutatenwechsel"], "existing": ["Kennzeichnung"], "measures":[M("Rein-/Unrein-Organisation"), M("Unterweisung LMIV","Q (Qualifikation/Unterweisung)")]},
        {"activity": "Elektrische Kleingeräte", "hazard": "Stromschlag, Brandrisiko", "sources": ["Mixer","Pürierstab"], "existing": ["Sichtprüfung"], "measures":[M("Prüfintervall ortsveränderliche Geräte")]},
    ],
    # ... (LIB_HOTEL restliche Bereiche bleiben wie in deiner Vorlage)
}

# --- Bäckerei (erweitert) ---
# (Hier steht deine zuletzt gepostete, erweiterte LIB_BAECKEREI – unverändert)
LIB_BAECKEREI = {  # ... kompletter Inhalt aus deiner letzten Version ...
    # (um Platz zu sparen: 1:1 aus deiner Nachricht übernehmen)
}

# --- Fleischerei / Fleischindustrie (erweitert: Schlachtung, Zerlegung, Verarbeitung) ---
LIB_FLEISCHEREI = {  # ... kompletter Inhalt aus deiner letzten Version ...
}

# --- Gemeinschaftsverpflegung / Kantine / Catering (stark erweitert) ---
LIB_KANTINE = {  # ... kompletter Inhalt aus deiner letzten Version ...
}

# --- Konditorei/Café, Brauerei, Getränkeabfüllung, Eisherstellung, Event, QSR, Wäscherei ---
LIB_KONDITOREI = {  # ... wie gepostet ...
}
LIB_BRAUEREI = {  # ... wie gepostet ...
}
LIB_GETRAENKEABF = {  # ... wie gepostet ...
}
LIB_EIS = {  # ... wie gepostet ...
}
LIB_EVENT = {  # ... wie gepostet ...
}
LIB_QSR = {  # ... wie gepostet ...
}
LIB_WAESCHE = {  # ... wie gepostet ...
}

INDUSTRY_LIBRARY: Dict[str, Dict[str, List[Dict[str, Any]]]] = {
    "Hotel/Gastgewerbe": LIB_HOTEL,
    "Bäckerei": LIB_BAECKEREI,
    "Fleischerei/Metzgerei": LIB_FLEISCHEREI,
    "Gemeinschaftsverpflegung/Kantine": LIB_KANTINE,
    "Konditorei/Café": LIB_KONDITOREI,
    "Brauerei": LIB_BRAUEREI,
    "Getränkeabfüllung": LIB_GETRAENKEABF,
    "Eisherstellung": LIB_EIS,
    "Event/Catering": LIB_EVENT,
    "Fast Food/Quickservice": LIB_QSR,
    "Wäscherei/Textilreinigung": LIB_WAESCHE,
}

# =========================
# Vorlagen laden/auswählen
# =========================

def add_template_items(
    assess: Assessment,
    template: Dict[str, List[Dict[str, Any]]],
    selected_keys: Optional[List[str]] = None,
    industry_name: Optional[str] = None,
    split_multi: Optional[bool] = None
):
    """Fügt Items aus einer Branchenvorlage hinzu (robust)."""
    if split_multi is None:
        split_multi = st.session_state.get("opt_split_multi_hazards", True)

    DEFAULT_STOP = "O (Organisatorisch)"

    def normalize_measure(m: Any) -> Optional[Measure]:
        if isinstance(m, dict):
            return Measure(
                title=(m.get("title") or "").strip(),
                stop_level=m.get("stop_level", DEFAULT_STOP),
                notes=m.get("notes", "")
            )
        elif isinstance(m, str):
            t = m.strip()
            return Measure(title=t, stop_level=DEFAULT_STOP) if t else None
        else:
            return None

    for area, items in template.items():
        for item in items:
            key = template_item_key(industry_name or assess.industry, area, item)
            if selected_keys is not None and key not in selected_keys:
                continue

            hazard_text = item.get("hazard", "")
            hazards_list = split_hazard_text(hazard_text) if split_multi else [hazard_text]

            for hz_text in hazards_list:
                hz = Hazard(
                    id=new_id(),
                    area=area,
                    activity=item.get("activity", ""),
                    hazard=hz_text,
                    sources=item.get("sources", []) or [],
                    existing_controls=item.get("existing", []) or []
                )
                for m in item.get("measures", []) or []:
                    mm = normalize_measure(m)
                    if mm and mm.title:
                        hz.additional_measures.append(mm)
                assess.hazards.append(hz)

def preload_industry(assess: Assessment, industry_name: str, replace: bool = True):
    assess.industry = industry_name
    if replace:
        assess.hazards = []
    template = INDUSTRY_LIBRARY.get(industry_name, {})
    add_template_items(assess, template, selected_keys=None, industry_name=industry_name)

def template_item_key(industry: str, area: str, item: Dict[str, Any]) -> str:
    return slug(industry, area, item.get("activity",""), item.get("hazard",""))

def iter_template_items(industry: str) -> List[Tuple[str, Dict[str, Any], str]]:
    lib = INDUSTRY_LIBRARY.get(industry, {})
    out = []
    for area, items in lib.items():
        for it in items:
            out.append((area, it, template_item_key(industry, area, it)))
    return out

# =========================
# Streamlit App (in main())
# =========================

def main():
    # Session init (robust)
    if "assessment" not in st.session_state or st.session_state.get("assessment") is None:
        st.session_state.assessment = Assessment(
            company="Musterbetrieb GmbH", location="Beispielstadt",
            created_at=date.today().isoformat(), created_by="HSE/SiFa",
            industry="Hotel/Gastgewerbe",
        )
        preload_industry(st.session_state.assessment, "Hotel/Gastgewerbe", replace=True)
    if "opt_split_multi_hazards" not in st.session_state:
        st.session_state["opt_split_multi_hazards"] = True
    if "json_blob" not in st.session_state:
        st.session_state["json_blob"] = ""

    assess: Assessment = st.session_state.assessment

    # Kopf
    col_head1, col_head2 = st.columns([0.8, 0.2])
    with col_head1:
        st.title("SiFa Toolkit – Alles für die Fachkraft für Arbeitssicherheit")
    with col_head2:
        if st.button("📄 Duplizieren", key="btn_duplicate"):
            assess.created_at = date.today().isoformat()
            assess.company = f"{assess.company} (Kopie)"
            st.success("Kopie erstellt. Bitte speichern/exportieren.")

    st.caption("Struktur: Vorlagen auswählen → Vorbereiten → Ermitteln → Beurteilen → Maßnahmen → Umsetzen → Wirksamkeit → Dokumentieren → Fortschreiben")

    # Sidebar
    with st.sidebar:
        st.header("Stammdaten")
        assess.company = st.text_input("Unternehmen", assess.company, key="meta_company")
        assess.location = st.text_input("Standort", assess.location, key="meta_location")
        assess.created_by = st.text_input("Erstellt von", assess.created_by, key="meta_created_by")
        assess.created_at = st.text_input("Erstellt am (ISO)", assess.created_at, key="meta_created_at")

        st.markdown("---")
        st.subheader("Branche wählen (für Vorlagen)")
        options = list(INDUSTRY_LIBRARY.keys())
        current_industry = getattr(assess, "industry", None) or "Hotel/Gastgewerbe"
        default_idx = options.index(current_industry) if current_industry in options else 0
        sector = st.selectbox("Branche", options=options, index=default_idx, key="sel_industry")
        st.caption(f"Aktuell geladen: **{assess.industry}**")

        # --- Optionen ---
        st.markdown("---")
        st.subheader("Optionen")
        if "opt_split_multi_hazards" not in st.session_state:
            st.session_state["opt_split_multi_hazards"] = True
        st.checkbox(
            "Mehrfach-Gefährdungen einer Tätigkeit automatisch auftrennen (1 Tätigkeit → 1 Gefährdung pro Eintrag)",
            key="opt_split_multi_hazards",
        )

        # Automatisches Nachladen bei Branchenwechsel
        st.markdown("---")
        st.caption("Automatisches Laden beim Branchenwechsel (optional)")
        if "last_sector" not in st.session_state:
            st.session_state.last_sector = sector
        elif st.session_state.last_sector != sector:
            assess.hazards = []
            tmpl = INDUSTRY_LIBRARY.get(sector, {})
            add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
            assess.industry = sector
            st.session_state.last_sector = sector
            st.toast(f"Vorlage '{sector}' automatisch geladen.", icon="✅")
            st.rerun()

        # Schnell-Laden
        st.markdown("---")
        st.markdown("**Schnell laden:**")
        c_load1, c_load2 = st.columns(2)
        with c_load1:
            if st.button("📚 Vorlage ERSETZEN", key="btn_load_replace_sidebar"):
                assess.hazards = []
                tmpl = INDUSTRY_LIBRARY.get(sector, {})
                add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
                assess.industry = sector
                if "template_checks" in st.session_state:
                    st.session_state.template_checks = {}
                st.success(f"Vorlage '{sector}' geladen (ersetzt).")
                st.rerun()
        with c_load2:
            if st.button("➕ Vorlage ANHÄNGEN", key="btn_load_append_sidebar"):
                tmpl = INDUSTRY_LIBRARY.get(sector, {})
                add_template_items(assess, tmpl, selected_keys=None, industry_name=sector)
                assess.industry = sector
                st.success(f"Vorlage '{sector}' hinzugefügt (angehängt).")
                st.rerun()

        st.markdown("---")
        st.subheader("Risikomatrix (5×5)")
        thr = assess.risk_matrix_thresholds.get("thresholds", [6, 12, 16])
        low = st.number_input("Grenze niedrig (≤)", min_value=2, max_value=10, value=int(thr[0]), key="thr_low")
        mid = st.number_input("Grenze mittel (≤)", min_value=low+1, max_value=16, value=int(thr[1]), key="thr_mid")
        high = st.number_input("Grenze hoch (≤)", min_value=mid+1, max_value=24, value=int(thr[2]), key="thr_high")
        assess.risk_matrix_thresholds["thresholds"] = [low, mid, high]

        st.markdown("---")
        st.subheader("Export / Speicher")
        if st.button("📥 JSON sichern (Download unten aktualisieren)", key="btn_json_dump"):
            st.session_state["json_blob"] = as_json(assess)
        json_blob = st.session_state.get("json_blob", as_json(assess))
        st.download_button("⬇️ Download JSON", data=json_blob, file_name="gefaehrdungsbeurteilung.json", mime="application/json", key="btn_dl_json")

        excel_bytes = dump_excel(assess)
        st.download_button("⬇️ Download Excel", data=excel_bytes, file_name="Gefaehrdungsbeurteilung.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="btn_dl_excel")

        st.markdown("---")
        st.subheader("JSON laden")
        up = st.file_uploader("Bestehende Beurteilung (.json)", type=["json"], key="uploader_json")
        if up is not None:
            content = up.read().decode("utf-8")
            st.session_state.assessment = from_json(content)
            if not getattr(st.session_state.assessment, "industry", None):
                st.session_state.assessment.industry = "Hotel/Gastgewerbe"
            st.success("Beurteilung geladen.")
            st.rerun()

    # Tabs
    tabs = st.tabs([
        "0 Vorlagen auswählen", "1 Vorbereiten", "2 Ermitteln", "3 Beurteilen", "4 Maßnahmen",
        "5 Umsetzen", "6 Wirksamkeit", "7 Dokumentation", "8 Fortschreiben", "Übersicht"
    ])

    # 0 Vorlagen auswählen
    with tabs[0]:
        st.subheader("0) Vorlagen auswählen (Tätigkeiten/Gefährdungen per Häkchen übernehmen)")
        st.caption("Branche wählen, filtern, Häkchen setzen, dann übernehmen. Mehrfach-Gefährdungen werden – wenn Option aktiv – automatisch in Einzel-Gefährdungen getrennt.")

        lib = INDUSTRY_LIBRARY.get(sector, {})
        all_areas = list(lib.keys())
        area_filter = st.multiselect("Bereiche filtern", options=all_areas, default=all_areas, key="tmpl_area_filter")
        text_filter = st.text_input("Textfilter (Activity/Gefährdung enthält…)", key="tmpl_text_filter").strip().lower()

        if "template_checks" not in st.session_state:
            st.session_state.template_checks = {}

        cols = st.columns([0.24, 0.24, 0.42, 0.10])
        cols[0].markdown("**Bereich**")
        cols[1].markdown("**Tätigkeit**")
        cols[2].markdown("**Gefährdung**")
        cols[3].markdown("**Auswählen**")

        items = iter_template_items(sector)
        shown_keys = []
        for area, item, keyval in items:
            if area_filter and area not in area_filter:
                continue
            if text_filter:
                blob = f"{item.get('activity','')} {item.get('hazard','')}".lower()
                if text_filter not in blob:
                    continue
            shown_keys.append(keyval)
            c0, c1, c2, c3 = st.columns([0.24, 0.24, 0.42, 0.10])
            c0.write(area)
            c1.write(item.get("activity",""))
            c2.write(item.get("hazard",""))
            default_checked = st.session_state.template_checks.get(keyval, False)
            st.session_state.template_checks[keyval] = c3.checkbox(" ", key=f"chk_{keyval}", value=default_checked)

        st.markdown("---")
        colA, colB, colC = st.columns([0.5,0.25,0.25])
        with colB:
            if st.button("Alle sichtbaren markieren", key="btn_mark_all"):
                for k in shown_keys:
                    st.session_state.template_checks[k] = True
                st.rerun()
        with colC:
            if st.button("Alle sichtbaren demarkieren", key="btn_unmark_all"):
                for k in shown_keys:
                    st.session_state.template_checks[k] = False
                st.rerun()

        st.markdown("---")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("➕ Ausgewählte übernehmen (ANHÄNGEN)", key="btn_apply_append"):
                selected = [k for k, v in st.session_state.template_checks.items() if v]
                add_template_items(assess, lib, selected_keys=selected, industry_name=sector)
                st.success(f"{len(selected)} Aktivitäten übernommen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
        with col2:
            if st.button("🧹 Ausgewählte übernehmen (ERSETZEN)", key="btn_apply_replace"):
                selected = [k for k, v in st.session_state.template_checks.items() if v]
                assess.hazards = []
                add_template_items(assess, lib, selected_keys=selected, industry_name=sector)
                assess.industry = sector
                st.success(f"Vorlage ersetzt. {len(selected)} Aktivitäten übernommen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
                st.rerun()

        st.markdown("---")
        if st.button("📦 Komplette Branchenvorlage übernehmen (ERSETZEN) – ohne Auswahl", key="btn_full_template_replace"):
            assess.hazards = []
            add_template_items(assess, lib, selected_keys=None, industry_name=sector)
            assess.industry = sector
            if "template_checks" in st.session_state:
                st.session_state.template_checks = {}
            st.success(f"Komplette Vorlage '{sector}' geladen (Mehrfach-Gefährdungen ggf. aufgetrennt).")
            st.rerun()

    # 1 Vorbereiten
    with tabs[1]:
        st.subheader("1) Vorbereiten")
        assess.industry = st.selectbox(
            "Branche der Beurteilung", options=list(INDUSTRY_LIBRARY.keys()),
            index=list(INDUSTRY_LIBRARY.keys()).index(assess.industry) if assess.industry in INDUSTRY_LIBRARY else 0,
            key="assess_industry"
        )
        assess.scope_note = st.text_area(
            "Umfang / Arbeitsbereiche / Beteiligte",
            value=assess.scope_note, height=140, key="scope_note"
        )
        st.info("Mit Tab „0 Vorlagen auswählen“ kannst du weitere Tätigkeiten/Gefährdungen anfügen.")

    # 2 Ermitteln
    with tabs[2]:
        st.subheader("2) Gefährdungen ermitteln")

        # Weicher Fallback statt st.stop()
        if not assess.hazards:
            st.warning("Noch keine Gefährdungen vorhanden. Nutze Tab 0 oder die Sidebar, um eine Branchenvorlage zu laden.")
            if st.button("🚀 Branchenvorlage jetzt laden und Beurteilung starten (ERSETZEN)", key="btn_fallback_load_from_tab2"):
                assess.hazards = []
                current_sector = st.session_state.get("sel_industry", assess.industry)
                tmpl = INDUSTRY_LIBRARY.get(current_sector, {})
                add_template_items(assess, tmpl, selected_keys=None, industry_name=current_sector)
                assess.industry = current_sector
                st.success(f"Vorlage '{assess.industry}' geladen. Du kannst jetzt beurteilen.")
                st.rerun()

                colL, colR = st.columns([2, 1])

        with colL:
            st.markdown("**Gefährdungen (Bearbeiten)**")
            if assess.hazards:
                df = pd.DataFrame([hazard_to_row(h) for h in assess.hazards])
                st.dataframe(df, use_container_width=True, hide_index=True, key="df_hazards")
            else:
                st.info("Nutze Tab 0 oder die Sidebar, um eine Branchenvorlage zu laden.")

            with st.expander("➕ Gefährdung manuell hinzufügen"):
                col1, col2 = st.columns(2)
                known_areas = sorted(
                    {h.area for h in assess.hazards}
                    | set(INDUSTRY_LIBRARY.get(assess.industry, {}).keys())
                    | {"Sonstiges"}
                )
                area = col1.selectbox("Bereich", known_areas, key="add_area")
                activity = col2.text_input("Tätigkeit", key="add_activity")
                hazard_txt = st.text_input(
                    "Gefährdung (bei mehreren: Komma/Slash/‚und‘ trennt in Einzeleinträge)",
                    key="add_hazard"
                )
                sources = st.text_input("Quellen/Einwirkungen (durch ; trennen)", key="add_sources")
                existing = st.text_input("Bestehende Maßnahmen (durch ; trennen)", key="add_existing")
                if st.button("Hinzufügen", key="btn_add_hazard"):
                    hazards_list = split_hazard_text(hazard_txt) if st.session_state.get("opt_split_multi_hazards", True) else [hazard_txt]
                    for hz_text in hazards_list:
                        assess.hazards.append(Hazard(
                            id=new_id(), area=area, activity=activity, hazard=hz_text,
                            sources=[s.strip() for s in sources.split(";") if s.strip()],
                            existing_controls=[e.strip() for e in existing.split(";") if e.strip()]
                        ))
                    st.success(f"{len(hazards_list)} Eintrag(e) hinzugefügt (1 Tätigkeit → 1 Gefährdung je Eintrag).")

        with colR:
            st.markdown("**Auswahl & Details**")
            ids = [h.id for h in assess.hazards]
            sel_id = st.selectbox("Gefährdung auswählen (ID)", options=["--"] + ids, index=0, key="sel_hazard_edit")
            if sel_id != "--":
                hz = next(h for h in assess.hazards if h.id == sel_id)
                all_areas = list(INDUSTRY_LIBRARY.get(assess.industry, {}).keys()) + ["Sonstiges"]
                idx = all_areas.index(hz.area) if hz.area in all_areas else len(all_areas) - 1
                hz.area = st.selectbox("Bereich", options=all_areas, index=idx, key=f"edit_area_{hz.id}")
                hz.activity = st.text_input("Tätigkeit", value=hz.activity, key=f"edit_activity_{hz.id}")
                hz.hazard = st.text_input("Gefährdung (nur eine pro Eintrag)", value=hz.hazard, key=f"edit_hazard_{hz.id}")
                src = st.text_area("Quellen/Einwirkungen", value="; ".join(hz.sources), key=f"edit_sources_{hz.id}")
                hz.sources = [s.strip() for s in src.split(";") if s.strip()]
                ex = st.text_area("Bestehende Maßnahmen", value="; ".join(hz.existing_controls), key=f"edit_existing_{hz.id}")
                hz.existing_controls = [e.strip() for e in ex.split(";") if e.strip()]
                if st.button("🗑️ Löschen", key=f"btn_delete_{hz.id}"):
                    assess.hazards = [h for h in assess.hazards if h.id != sel_id]
                    st.warning("Gefährdung gelöscht.")
                    st.rerun()
